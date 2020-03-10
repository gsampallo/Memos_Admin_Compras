from openpyxl import load_workbook
import os
import shutil

wb = load_workbook(filename = '/Users/jbparra/Downloads/Memo. Casino 0001.xlsx')

sheet = wb.active

contador = 0
memo_nro = sheet['B2'].value
fecha_memo = sheet['G2'].value
print('Memo Nro: ' + memo_nro)
print('Fecha Memo: ' + str(fecha_memo))

# Loop que busca en las filas de proveedor, factura y fecha
for value in sheet.iter_rows(min_row = 5, max_row = 10, min_col = 1, max_col = 3, values_only = True):

    # chequea que la fila tenga valores
    if not value[0] and contador == 0:
        print('No se encuentran valores')
        break
    if not value[0] and contador > 0:
        print('No se encuentran mas valores')
        break
    else:
        proveedor = value[0]
        factura = value[1]
        fecha = value[2].strftime("%Y%m%d")

        contador =+ 1
        
        print('Contador: ' + str(contador))
        print('Proveedor: ' + proveedor + ' - Factura: ' + str(factura) + ' - Fecha: ' + str(fecha))

        nombre_archivo_scan = fecha + '_' + proveedor + '_' + str(factura)

        # Renombre del archivo de scan
        os.rename('/Users/jbparra/Downloads/statements.pdf', '/Users/jbparra/Downloads/' + nombre_archivo_scan + '.pdf')

        print(nombre_archivo_scan)

# Crea carpeta Memo con Nro identificatorio
path = os.getcwd()
os.mkdir(path + '/Memo Nro ' + memo_nro)

source = '/Users/jbparra/Desktop/test/'
destination = '/Users/jbparra/Desktop/test2/'

# Busca todos los archivos en el directorio
files = os.listdir(source)

# Mueve los archivos con un loop
for f in files:
    shutil.move(source+f, destination)
